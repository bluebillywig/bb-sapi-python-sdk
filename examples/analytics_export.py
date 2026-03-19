"""
Analytics export example — top videos with ad stats.

Generates an Excel workbook with:
  - Sheet 1: Top N content videos by views (with viewcount reach & ad metrics)
  - Sheet 2: Per-video pre-roll creative breakdown

Usage::

    pip install bb-sapi-python-sdk openpyxl
    python examples/analytics_export.py

Configure via environment variables or edit the CONFIG dict below.
"""
from __future__ import annotations

import os
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("Install openpyxl: pip install openpyxl")

from bb_sapi import SapiClient
from bb_sapi.exceptions import SapiAuthError, SapiError

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
CONFIG = {
    "base_url": os.getenv("SAPI_BASE_URL", "https://mypublication.bbvms.com"),
    "shared_secret": os.getenv("SAPI_SHARED_SECRET", ""),
    "from_date": os.getenv("SAPI_FROM_DATE", "2026-01-01"),
    "to_date": os.getenv("SAPI_TO_DATE", "2026-03-31"),
    "top_n": int(os.getenv("SAPI_TOP_N", "50")),
    "output_file": os.getenv("SAPI_OUTPUT", "analytics_export.xlsx"),
}

REACH_THRESHOLDS = (20, 40, 60, 80, 95)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def fetch_top_videos(client: SapiClient, cfg: dict[str, Any]) -> list[dict[str, Any]]:
    """Fetch the top N content videos by view count."""
    top = client.analytics.top_videos(cfg["from_date"], cfg["to_date"], limit=cfg["top_n"])
    # Enrich with metadata and filter out ad creatives (usetype: commercial)
    ids = [v["id"] for v in top]
    meta: dict[str, dict] = {}
    for chunk_ids in _chunks(ids, 50):
        q = "id:(" + " OR ".join(str(i) for i in chunk_ids) + ")"
        results = client.search(q, entity_type="MediaClip", fields="id,title,usetype,duration")
        for item in results.get("items", []):
            meta[str(item["id"])] = item

    content = []
    for v in top:
        vid_meta = meta.get(str(v["id"]))
        if vid_meta is None:
            print(f"  Warning: video {v['id']} not found in metadata search; skipping.")
            continue
        if vid_meta.get("usetype") != "commercial":
            v.update(vid_meta)
            content.append(v)
    return content


def enrich_with_stats(
    client: SapiClient,
    videos: list[dict[str, Any]],
    cfg: dict[str, Any],
) -> None:
    """Add analytics stats in-place to each video dict."""
    from_date, to_date = cfg["from_date"], cfg["to_date"]
    failed = 0
    for v in videos:
        vid_id = str(v["id"])
        try:
            ad = client.analytics.ad_stats_per_video(vid_id, from_date, to_date)
            reach = client.analytics.viewcount_reach(
                vid_id, from_date, to_date, thresholds=REACH_THRESHOLDS
            )
        except SapiAuthError:
            raise  # credentials broken — abort immediately
        except SapiError as exc:
            print(
                f"  ERROR: failed stats for {vid_id} ({v.get('title', '?')}): {exc}; "
                f"recording as empty."
            )
            v["impressions"] = None
            v["lineitems"] = {}
            v["vast_quartiles"] = {}
            v["reach"] = {}
            failed += 1
            continue

        v["impressions"] = ad["impressions"]
        v["lineitems"] = ad["lineitems"]
        v["vast_quartiles"] = ad["vastQuartiles"]
        v["reach"] = reach
        print(f"  {v.get('title', vid_id)}: {v['views']} views, {v['impressions']} impressions")

    if failed:
        print(f"  WARNING: {failed}/{len(videos)} videos failed stats fetch.")


def fetch_creatives(
    client: SapiClient,
    videos: list[dict[str, Any]],
    cfg: dict[str, Any],
) -> dict[str, list[dict]]:
    """For each lineitem that appeared, fetch which creative(s) were active."""
    all_lineitems: set[str] = set()
    for v in videos:
        all_lineitems.update(v.get("lineitems", {}).keys())

    creative_map: dict[str, list[dict]] = {}
    failed = 0
    for name in sorted(all_lineitems):
        try:
            creative_map[name] = client.lineitem.creatives_for_period(
                name, cfg["from_date"], cfg["to_date"]
            )
        except SapiAuthError:
            raise  # credentials broken — abort immediately
        except SapiError as exc:
            print(f"  Warning: could not fetch creatives for {name!r}: {exc}")
            creative_map[name] = []
            failed += 1

    if failed:
        print(f"  WARNING: {failed}/{len(all_lineitems)} lineitems failed creative resolution.")
    return creative_map


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_FILL = PatternFill("solid", fgColor="EBF3FB")


def _header(ws: Any, row: int, values: list[str]) -> None:
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def _autofit(ws: Any) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)


def build_workbook(
    videos: list[dict[str, Any]],
    creative_map: dict[str, list[dict]],
) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------
    # Sheet 1: video overview
    # ------------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Videos"
    headers = [
        "ID", "Title", "Views", "Impressions",
        "VAST 25%", "VAST 50%", "VAST 75%", "VAST 100%",
    ] + [f"Reach ≥{t}%" for t in REACH_THRESHOLDS]
    _header(ws1, 1, headers)

    for i, v in enumerate(videos, start=2):
        vq = v.get("vast_quartiles", {})
        reach = v.get("reach", {})
        row = [
            v["id"], v.get("title", ""), v["views"], v.get("impressions", 0),
            vq.get("25", 0), vq.get("50", 0), vq.get("75", 0), vq.get("100", 0),
        ] + [reach.get(t, 0) for t in REACH_THRESHOLDS]
        for col, val in enumerate(row, start=1):
            cell = ws1.cell(row=i, column=col, value=val)
            if i % 2 == 0:
                cell.fill = ALT_FILL
    _autofit(ws1)

    # ------------------------------------------------------------------
    # Sheet 2: pre-roll creative breakdown
    # ------------------------------------------------------------------
    ws2 = wb.create_sheet("Pre-roll Creatives")
    _header(ws2, 1, ["Lineitem", "Creative ID", "Version Date", "VAST URL"])
    row_num = 2
    for name in sorted(creative_map):
        for c in creative_map[name]:
            ws2.cell(row=row_num, column=1, value=name)
            ws2.cell(row=row_num, column=2, value=c.get("creative_id", "external"))
            ws2.cell(row=row_num, column=3, value=c.get("date", ""))
            ws2.cell(row=row_num, column=4, value=c.get("vast_url", ""))
            if row_num % 2 == 0:
                for col in range(1, 5):
                    ws2.cell(row=row_num, column=col).fill = ALT_FILL
            row_num += 1
    _autofit(ws2)

    return wb


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def _chunks(lst: list, size: int):
    for i in range(0, len(lst), size):
        yield lst[i : i + size]


def main() -> None:
    cfg = CONFIG
    if not cfg["shared_secret"]:
        raise SystemExit("Set SAPI_SHARED_SECRET environment variable.")

    client = SapiClient(cfg["base_url"], cfg["shared_secret"])

    print(f"Fetching top {cfg['top_n']} videos ({cfg['from_date']} – {cfg['to_date']})…")
    videos = fetch_top_videos(client, cfg)
    print(f"Found {len(videos)} content videos.")

    print("Fetching per-video stats…")
    enrich_with_stats(client, videos, cfg)

    print("Resolving pre-roll creatives…")
    creative_map = fetch_creatives(client, videos, cfg)

    print("Building workbook…")
    wb = build_workbook(videos, creative_map)
    wb.save(cfg["output_file"])
    print(f"Saved: {cfg['output_file']}")


if __name__ == "__main__":
    main()
